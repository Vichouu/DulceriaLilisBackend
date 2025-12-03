from datetime import datetime
import json
import re

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.views.decorators.http import require_POST

from lilis_erp.roles import require_roles

# Modelos
from apps.suppliers.models import Proveedor, ProveedorProducto
from apps.products.models import Producto

# Excel
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


# -------------------------- Helpers --------------------------

def _valid_email(s: str) -> bool:
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", s or ""))

# NUEVO: helpers para RUT chileno
def normalizar_rut(rut: str) -> str:
    """
    Elimina puntos y guión, deja sólo dígitos + K en mayúscula.
    """
    return (rut or "").replace(".", "").replace("-", "").strip().upper()

def rut_chileno_valido(rut: str) -> bool:
    """
    Valida RUT chileno:
      - 7 u 8 dígitos + DV (0-9 o K)
      - Usa algoritmo módulo 11
    """
    limpio = normalizar_rut(rut)
    if not limpio or not re.fullmatch(r"\d{7,8}[0-9K]", limpio):
        return False

    cuerpo = limpio[:-1]
    dv = limpio[-1]

    try:
        int(cuerpo)
    except ValueError:
        return False

    suma = 0
    multiplo = 2
    for d in reversed(cuerpo):
        suma += int(d) * multiplo
        multiplo += 1
        if multiplo > 7:
            multiplo = 2

    resto = 11 - (suma % 11)
    if resto == 11:
        dv_calc = "0"
    elif resto == 10:
        dv_calc = "K"
    else:
        dv_calc = str(resto)

    return dv == dv_calc

def _estado_from_text(q: str):
    """
    Normaliza a los estados guardados en BD (minúscula para matchear consistentes).
    """
    q = (q or '').strip().lower()
    mapping = {
        'activo': 'activo',
        'inactivo': 'inactivo',
        'bloqueado': 'bloqueado',
    }
    return mapping.get(q)

def _build_supplier_q(q: str) -> Q:
    """
    Búsqueda para proveedores:
    - rut_nif, razon_social, estado, email (requerido)
    - extras no invasivos (ayudan)
    """
    q = (q or "").strip()
    if not q:
        return Q()

    expr = (
        Q(rut_nif__icontains=q) |
        Q(razon_social__icontains=q) |
        Q(email__icontains=q) |
        # extras útiles
        Q(nombre_fantasia__icontains=q) |
        Q(telefono__icontains=q) |
        Q(direccion__icontains=q) |
        Q(ciudad__icontains=q) |
        Q(pais__icontains=q) |
        Q(contacto_principal_nombre__icontains=q) |
        Q(contacto_principal_email__icontains=q) |
        Q(contacto_principal_telefono__icontains=q) |
        Q(condiciones_pago__icontains=q)
    )

    # ID exacto si es número
    try:
        expr |= Q(id=int(q))
    except ValueError:
        pass

    return expr

def _build_relation_q(q: str) -> Q:
    q = (q or "").strip()
    if not q:
        return Q()
    return (
        Q(proveedor__rut_nif__icontains=q) |
        Q(proveedor__razon_social__icontains=q) |
        Q(producto__sku__icontains=q) |
        Q(producto__nombre__icontains=q)
    )


# ---------------------------- Vistas ----------------------------

@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
def supplier_list_view(request):
    query = request.GET.get('q', '')
    sort_by = request.GET.get('sort', 'id')
    ver = request.GET.get('ver', 'todos')
    export = request.GET.get('export', '')

    valid_sort_fields = ['id', '-id', 'rut_nif', '-rut_nif', 'razon_social', '-razon_social']
    if sort_by not in valid_sort_fields:
        sort_by = 'id'

    # Primero, filtramos por estado (ver)
    if ver == 'activos':
        qs = Proveedor.objects.filter(activo=True)
    elif ver == 'inactivos':
        qs = Proveedor.objects.filter(activo=False)
    else:  # 'todos'
        qs = Proveedor.objects.all()

    # Luego, sobre el resultado anterior, aplicamos la búsqueda por texto si existe
    if query:
        qs = qs.filter(_build_supplier_q(query))

    qs = qs.order_by(sort_by)

    # Export a Excel
    if export == "xlsx":
        if Workbook is None:
            return HttpResponse(
                "Falta dependencia: instala openpyxl (pip install openpyxl)",
                status=500,
                content_type="text/plain; charset=utf-8",
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Proveedores"
        headers = [
            "ID", "RUT/NIF", "Razón Social", "Nombre Fantasía", "Email", "Teléfono",
            "Sitio Web", "Plazos Pago (días)", "Moneda", "Descuento (%)", "Estado", "Activo"
        ]
        ws.append(headers)

        for p in qs:
            ws.append([
                p.id,
                getattr(p, "rut_nif", ""),
                getattr(p, "razon_social", ""),
                getattr(p, "nombre_fantasia", ""),
                getattr(p, "email", ""),
                getattr(p, "telefono", ""),
                getattr(p, "sitio_web", ""),
                getattr(p, "plazos_pago_dias", 0),
                getattr(p, "moneda", ""),
                getattr(p, "descuento_porcentaje", 0),
                getattr(p, "estado", ""),
                "Sí" if getattr(p, "activo", False) else "No",
            ])

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"proveedores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    # Relaciones proveedor-producto (filtradas también por q)
    relations_qs = ProveedorProducto.objects.select_related('proveedor', 'producto').order_by('-id')
    if query:
        relations_qs = relations_qs.filter(_build_relation_q(query))
    relations_paginator = Paginator(relations_qs, 10)
    relations_page_obj = relations_paginator.get_page(request.GET.get("page_rel"))

    # Normaliza estado para visual
    proveedores_list = []
    for p in page_obj.object_list:
        # Creamos un atributo 'estado_display' para no sobreescribir el original
        p.estado_display = "Activo" if p.activo else "Inactivo"
        proveedores_list.append(p)

    context = {
        "proveedores": proveedores_list,
        "page_obj": page_obj,
        "query": query,
        "sort_by": sort_by,
        "ver": ver,
        "relations": relations_page_obj,
    }
    return render(request, "gestion_proveedores.html", context)


# ------------------------- AJAX: CREATE -------------------------

@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
@require_POST
def create_supplier(request):
    """
    Crea/actualiza proveedor por RUT. Si existe, actualiza datos básicos.
    Validación:
      - rut_nif requerido y válido (RUT chileno) + ejemplo
      - razon_social requerida
      - email requerido y válido
    """
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "errors": {"__all__": "Payload inválido"}}, status=400)

    errors = {}
    rut = (data.get("rut_nif") or "").strip()
    razon = (data.get("razon_social") or "").strip()
    email = (data.get("email") or "").strip()
    telefono = (data.get("telefono") or "").strip()
    web = (data.get("sitio_web") or "").strip()
    nombre_fantasia = (data.get("nombre_fantasia") or "").strip()

    # comerciales
    condiciones_pago = (data.get("condiciones_pago") or "").strip()
    moneda = (data.get("moneda") or "CLP").strip()
    try:
        descuento = float(data.get("descuento_porcentaje") or 0)
    except Exception:
        descuento = -1

    # Validaciones
    if not rut_chileno_valido(rut):
        errors["rut_nif"] = "RUT/NIF obligatorio y válido. Ejemplo: 12.345.678-9."

    if not razon:
        errors["razon_social"] = "Razón social obligatoria."

    if not _valid_email(email):
        errors["email"] = "Email obligatorio y válido."

    if not telefono or not re.match(r"^[0-9+()\-\s]{6,30}$", telefono):
        errors["telefono"] = "Teléfono obligatorio y válido."

    if web and not re.match(r"^https?://.+", web, flags=re.I):
        errors["sitio_web"] = "Debe comenzar con http:// o https://"

    if not condiciones_pago:
        errors["condiciones_pago"] = "Las condiciones de pago son obligatorias."

    if not (0 <= descuento <= 100):
        errors["descuento_porcentaje"] = "Descuento 0 a 100%."

    # Duplicado por RUT exacto
    if not errors:
        existente = Proveedor.objects.filter(rut_nif=rut).first()
        if existente and str(existente.id) != str(data.get("id") or ""):
            errors["rut_nif"] = "Ya existe un proveedor con este RUT/NIF."

    if errors:
        return JsonResponse({"ok": False, "errors": errors}, status=400)

    # Crear o actualizar
    with transaction.atomic():
        proveedor, _created = Proveedor.objects.update_or_create(
            rut_nif=rut,
            defaults={
                "razon_social": razon,
                "email": email,
                "telefono": telefono,
                "sitio_web": web,
                "nombre_fantasia": nombre_fantasia,
                "condiciones_pago": condiciones_pago,
                "moneda": moneda or "CLP",
                "activo": True,
            }
        )

    return JsonResponse({"ok": True, "id": proveedor.id})


@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
@require_POST
def create_relation(request):
    """
    Crea la relación proveedor–producto:
      - requiere rut_nif de proveedor existente
      - busca producto por SKU exacto o por nombre (primer match)
      - valida no negativos y rangos
      - si ya existe la relación para el par (proveedor, producto), actualiza
    Campos aceptados:
      rut_nif, sku_or_name, preferente(bool), lead_time_dias(int 0-365),
      costo(float>=0), minimo_lote(int>=0 opc), descuento_porcentaje(0-100)
    """
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "errors": {"__all__": "Payload inválido"}}, status=400)

    errors = {}

    rut = (data.get("rut_nif") or "").strip()
    sku_or_name = (data.get("sku_or_name") or "").strip()
    preferente = bool(data.get("preferente"))
    try:
        lead = int(data.get("lead_time_dias") or 0)
    except Exception:
        lead = -1
    try:
        costo = float(data.get("costo") or 0)
    except Exception:
        costo = -1
    try:
        minimo_lote = int(data.get("minimo_lote") or 0)
    except Exception:
        minimo_lote = -1
    try:
        desc = float(data.get("descuento_porcentaje") or 0)
    except Exception:
        desc = -1

    if not rut:
        errors["rut_nif"] = "Debe indicar el RUT/NIF del proveedor (pestaña 1)."
    if not sku_or_name:
        errors["sku_or_name"] = "Ingrese SKU o nombre de producto."
    if not (0 <= lead <= 365):
        errors["lead_time_dias"] = "Lead time 0 a 365."
    if not (costo >= 0):
        errors["costo"] = "Costo no puede ser negativo."
    if not (minimo_lote >= 0):
        errors["minimo_lote"] = "Mínimo lote no puede ser negativo."
    if not (0 <= desc <= 100):
        errors["descuento_porcentaje"] = "Descuento 0 a 100%."

    proveedor = None
    producto = None

    if not errors:
        proveedor = Proveedor.objects.filter(rut_nif=rut).first()
        if not proveedor:
            errors["rut_nif"] = "Proveedor no existe. Guárdalo primero."

    if not errors:
        # buscar producto por SKU exacto o por nombre (primer resultado)
        producto = (Producto.objects.filter(sku=sku_or_name).first() or
                    Producto.objects.filter(nombre__icontains=sku_or_name).order_by("id").first())
        if not producto:
            errors["sku_or_name"] = "Producto no encontrado."

    if errors:
        return JsonResponse({"ok": False, "errors": errors}, status=400)

    with transaction.atomic():
        rel, _created = ProveedorProducto.objects.update_or_create(
            proveedor=proveedor,
            producto=producto,
            defaults={
                "preferente": preferente,
                "lead_time_dias": lead,
                "costo": costo,
                "minimo_lote": minimo_lote,
                "descuento_porcentaje": desc,
            }
        )

    return JsonResponse({"ok": True, "id": rel.id})


# --------------------- AJAX: SEARCH / EXPORT ---------------------

@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
def search_suppliers(request):
    q = (request.GET.get("q") or "").strip()
    qs = Proveedor.objects.all()
    if q:
        qs = qs.filter(_build_supplier_q(q))
    qs = qs.order_by("id")[:10]
    return JsonResponse({"results": [{
        "id": s.id,
        "rut": getattr(s, "rut_nif", ""),
        "razon_social": getattr(s, "razon_social", ""),
        "estado": (s.estado or "").capitalize() if s.estado else ("Activo" if getattr(s, "activo", False) else "Inactivo"),
        "email": getattr(s, "email", ""),
    } for s in qs]})

@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
def relations_search(request):
    q = (request.GET.get("q") or "").strip()
    qs = ProveedorProducto.objects.select_related("proveedor", "producto")
    if q:
        qs = qs.filter(_build_relation_q(q))
    qs = qs.order_by("id")[:10]

    def _rel_to_dict(rel: ProveedorProducto):
        pvd = rel.proveedor
        pro = rel.producto
        return {
            "id": rel.id,
            "proveedor": getattr(pvd, "razon_social", ""),
            "rut": getattr(pvd, "rut_nif", ""),
            "producto": getattr(pro, "nombre", ""),
            "sku": getattr(pro, "sku", ""),
            "preferente": bool(getattr(rel, "preferente", False)),
            "lead_time": getattr(rel, "lead_time_dias", 0) or 0,
            "costo": getattr(rel, "costo", 0) or 0,
            "minimo_lote": getattr(rel, "minimo_lote", 0) or 0,
            "descuento_porcentaje": getattr(rel, "descuento_porcentaje", 0) or 0,
        }

    return JsonResponse({"results": [_rel_to_dict(r) for r in qs]})

@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
def relations_export(request):
    if Workbook is None:
        return HttpResponse(
            "Falta dependencia: instala openpyxl (pip install openpyxl)",
            status=500,
            content_type="text/plain; charset=utf-8",
        )

    q = (request.GET.get("q") or "").strip()
    qs = ProveedorProducto.objects.select_related("proveedor", "producto")
    if q:
        qs = qs.filter(_build_relation_q(q))
    qs = qs.order_by("id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Relaciones"
    headers = ["ID", "Proveedor", "RUT/NIF", "Producto", "SKU",
               "Preferente", "Lead time (d)", "Costo", "Mínimo lote", "Descuento (%)"]
    ws.append(headers)

    for r in qs:
        pvd = r.proveedor
        pro = r.producto
        ws.append([
            r.id,
            getattr(pvd, "razon_social", ""),
            getattr(pvd, "rut_nif", ""),
            getattr(pro, "nombre", ""),
            getattr(pro, "sku", ""),
            "Sí" if getattr(r, "preferente", False) else "No",
            getattr(r, "lead_time_dias", 0) or 0,
            getattr(r, "costo", 0) or 0,
            getattr(r, "minimo_lote", 0) or 0,
            getattr(r, "descuento_porcentaje", 0) or 0,
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"relaciones_proveedor_producto_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ---------------------- EDITAR / ESTADO / ELIMINAR ----------------------

@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
def editar_proveedor(request, supplier_id):
    try:
        proveedor = Proveedor.objects.get(id=supplier_id)
    except Proveedor.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Proveedor no encontrado"}, status=404)

    if request.method == "GET":
        return JsonResponse({
            "id": proveedor.id,
            "rut_nif": proveedor.rut_nif,
            "razon_social": proveedor.razon_social,
            "email": proveedor.email,
            "telefono": proveedor.telefono or "",
            "sitio_web": proveedor.sitio_web or "",
            "condiciones_pago": proveedor.condiciones_pago or "",
            "moneda": proveedor.moneda or "CLP",
        })

    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            proveedor.rut_nif = data.get("rut_nif", proveedor.rut_nif).strip()
            proveedor.razon_social = data.get("razon_social", proveedor.razon_social).strip()
            proveedor.email = data.get("email", proveedor.email).strip()
            proveedor.telefono = data.get("telefono", proveedor.telefono).strip()
            proveedor.sitio_web = data.get("sitio_web", proveedor.sitio_web).strip()
            proveedor.condiciones_pago = data.get("condiciones_pago", proveedor.condiciones_pago).strip()

            # NUEVO: validar RUT al editar también
            if not rut_chileno_valido(proveedor.rut_nif):
                return JsonResponse({
                    "status": "error",
                    "message": "RUT/NIF inválido. Ejemplo: 12.345.678-9."
                }, status=400)

            if Proveedor.objects.filter(rut_nif=proveedor.rut_nif).exclude(id=supplier_id).exists():
                return JsonResponse({"status": "error", "message": "Ya existe otro proveedor con este RUT/NIF."}, status=400)

            proveedor.save()
            return JsonResponse({"status": "ok", "message": "Proveedor actualizado correctamente"})
        except json.JSONDecodeError:
            return JsonResponse({"status": "error", "message": "JSON inválido"}, status=400)
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
@require_POST
def eliminar_proveedor(request, supplier_id):
    try:
        with transaction.atomic():
            proveedor = Proveedor.objects.get(id=supplier_id)
            # Primero, eliminamos todas las relaciones que este proveedor pueda tener
            ProveedorProducto.objects.filter(proveedor=proveedor).delete()
            # Ahora sí, eliminamos el proveedor
            proveedor.delete()
            return JsonResponse({"status": "ok", "message": "Proveedor y sus relaciones eliminados correctamente"})
    except Proveedor.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Proveedor no encontrado"}, status=404)
    except Exception as e:
        # Captura cualquier otro error de base de datos o de lógica
        return JsonResponse({"status": "error", "message": f"No se pudo eliminar el proveedor. Error: {str(e)}"}, status=500)

@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
@require_POST
def desactivar_proveedor(request, supplier_id):
    try:
        proveedor = Proveedor.objects.get(id=supplier_id)
    except Proveedor.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Proveedor no encontrado"}, status=404)

    try:
        proveedor.estado = 'inactivo'
    except Exception:
        pass
    proveedor.activo = False
    proveedor.save(update_fields=[f for f in ['estado', 'activo'] if hasattr(proveedor, f)])
    return JsonResponse({'status': 'ok', 'message': 'Proveedor desactivado.'})

@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
@require_POST
def reactivar_proveedor(request, supplier_id):
    try:
        proveedor = Proveedor.objects.get(id=supplier_id)
    except Proveedor.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Proveedor no encontrado"}, status=404)

    try:
        proveedor.estado = 'activo'
    except Exception:
        pass
    proveedor.activo = True
    proveedor.save(update_fields=[f for f in ['estado', 'activo'] if hasattr(proveedor, f)])
    return JsonResponse({'status': 'ok', 'message': 'Proveedor reactivado.'})

@login_required
@require_roles("ADMIN", "COMPRAS", "INVENTARIO")
@require_POST
def eliminar_relacion(request, relation_id):
    try:
        relacion = ProveedorProducto.objects.get(id=relation_id)
        relacion.delete()
        return JsonResponse({"status": "ok", "message": "Relación eliminada correctamente"})
    except ProveedorProducto.DoesNotExist:
        return JsonResponse({"status": "error", "message": "La relación no fue encontrada."}, status=404)
    except Exception as e:
        return JsonResponse({"status": "error", "message": f"No se pudo eliminar la relación. Error: {str(e)}"}, status=500)
