from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden, HttpRequest
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db.models import Q
from django.db import transaction
from datetime import datetime
from apps.account.views import get_redirect_for_role
from .models import Usuario
from .utils_invite import invite_user_and_email
from .forms import UsuarioForm

# ====== AUDITORÍA ======
import logging
audit_logger = logging.getLogger("auditoria")
# =======================

# ====== export a Excel (openpyxl) ======
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


def _usuarios_to_excel(queryset):
    if Workbook is None:
        return HttpResponse(
            "Falta dependencia: instala openpyxl (pip install openpyxl)",
            status=500,
            content_type="text/plain; charset=utf-8",
        )
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    headers = ["ID", "Username", "Email", "Nombre", "Apellido", "Teléfono",
               "Rol", "Estado", "Activo", "MFA", "Último acceso", "Creado"]
    ws.append(headers)
    for u in queryset:
        ws.append([
            u.id, u.username, u.email, u.first_name or "", u.last_name or "",
            u.telefono or "", getattr(u, "rol", ""), getattr(u, "estado", ""),
            "Sí" if u.activo else "No",
            "Sí" if getattr(u, "mfa_habilitado", False) else "No",
            u.last_login.strftime("%d/%m/%Y %H:%M") if u.last_login else "",
            u.date_joined.strftime("%d/%m/%Y %H:%M") if u.date_joined else "",
        ])
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"usuarios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _rol_from_text(q: str):
    q = (q or "").strip().lower()
    mapping = {
        'administrador': 'ADMIN',
        'admin': 'ADMIN',
        'compras': 'COMPRAS',
        'inventario': 'INVENTARIO',
        'produccion': 'PRODUCCION',
        'producción': 'PRODUCCION',
        'ventas': 'VENTAS',
        'finanzas': 'FINANZAS',
        'soporte': 'SOPORTE',
    }
    return mapping.get(q)


def _es_admin(user):
    return user.is_superuser or getattr(user, "rol", "") == "ADMIN"


@login_required
def gestion_usuarios(request):
    if not _es_admin(request.user):
        return redirect(get_redirect_for_role(request.user))

    query = request.GET.get('q', '')
    sort_by = request.GET.get('sort', 'id')
    export = request.GET.get('export')
    ver = request.GET.get('ver', 'todos')

    valid_sort_fields = ['id', '-id', 'username', '-username', 'first_name', '-first_name', 'rol', '-rol']
    if sort_by not in valid_sort_fields:
        sort_by = 'id'

    usuarios_list = Usuario.objects.all()

    if ver == 'inactivos':
        usuarios_list = usuarios_list.filter(Q(estado__in=['inactivo', 'bloqueado']) | Q(activo=False))
    elif ver == 'activos':
        usuarios_list = usuarios_list.filter(estado='activo', activo=True)

    q = (query or '').strip()
    expr = Q()
    if q:
        expr |= Q(username__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q)
        expr |= Q(email__icontains=q) | Q(telefono__icontains=q)
        try:
            expr |= Q(id=int(q))
        except ValueError:
            pass
        rol_code = _rol_from_text(q)
        if rol_code:
            expr |= Q(rol=rol_code)
        usuarios_list = usuarios_list.filter(expr)

    usuarios_list = usuarios_list.order_by(sort_by)

    if export == 'xlsx':
        return _usuarios_to_excel(usuarios_list)

    paginator = Paginator(usuarios_list, 10)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'usuarios/gestion_usuarios.html', {
        'page_obj': page_obj,
        'query': query,
        'sort_by': sort_by,
        'ver': ver
    })


# ================================================================
# 🔥 CRUD + Auditoría
# ================================================================

@login_required
@transaction.atomic
def crear_usuario(request):
    if not _es_admin(request.user):
        return HttpResponseForbidden("Solo Administrador.")

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

    form = UsuarioForm(request.POST)
    if form.is_valid():
        usuario = form.save(commit=False)
        usuario.activo = (form.cleaned_data.get('estado') == 'activo')
        usuario.set_unusable_password()
        usuario.save()

        invite_user_and_email(usuario)

        # === AUDITORÍA ===
        audit_logger.info(
            f"CREATE Usuario id={usuario.id}, username={usuario.username}, por={request.user.username}"
        )

        return JsonResponse({'status': 'ok', 'message': 'Usuario creado e invitación enviada.'})

    return JsonResponse({'status': 'error', 'errors': form.errors.get_json_data()}, status=400)



@login_required
@csrf_exempt
def eliminar_usuario(request, user_id):
    if not _es_admin(request.user):
        return HttpResponseForbidden("Solo Administrador.")

    usuario = get_object_or_404(Usuario, id=user_id)

    if usuario.id == request.user.id:
        return JsonResponse({'status': 'error', 'message': 'No puedes eliminar tu propia cuenta.'}, status=403)

    if _es_admin(usuario) and not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Solo un Superusuario puede eliminar a otro administrador.'}, status=403)

    # === AUDITORÍA ===
    audit_logger.info(
        f"DELETE Usuario id={usuario.id}, username={usuario.username}, por={request.user.username}"
    )

    usuario.delete()
    return JsonResponse({'status': 'ok', 'message': 'Usuario eliminado correctamente.'})



@login_required
@csrf_exempt
def editar_usuario(request, user_id):
    if not _es_admin(request.user):
        return HttpResponseForbidden("Solo Administrador.")

    usuario = get_object_or_404(Usuario, id=user_id)

    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            usuario_actualizado = form.save(commit=False)
            usuario_actualizado.activo = (form.cleaned_data.get('estado') == 'activo')
            usuario_actualizado.save()

            # === AUDITORÍA ===
            audit_logger.info(
                f"UPDATE Usuario id={usuario.id}, username={usuario.username}, por={request.user.username}"
            )

            return JsonResponse({'status': 'ok', 'message': 'Usuario actualizado correctamente.'})

        return JsonResponse({'status': 'error', 'errors': form.errors.get_json_data()}, status=400)

    return JsonResponse({
        'id': usuario.id,
        'username': usuario.username,
        'email': usuario.email,
        'first_name': usuario.first_name,
        'last_name': usuario.last_name,
        'telefono': usuario.telefono,
        'rol': usuario.rol,
        'estado': usuario.estado,
        'mfa_habilitado': usuario.mfa_habilitado,
    })



# ================================================================
# 🔥 Activación / desactivación + Auditoría
# ================================================================

@login_required
@csrf_exempt
def desactivar_usuario(request, user_id):
    if not _es_admin(request.user):
        return HttpResponseForbidden("Solo Administrador.")
    usuario = get_object_or_404(Usuario, id=user_id)

    if usuario.id == request.user.id:
        return JsonResponse({'status': 'error', 'message': 'No puedes desactivar tu propia cuenta.'}, status=403)

    usuario.estado = 'inactivo'
    usuario.activo = False
    usuario.save(update_fields=['estado', 'activo'])

    audit_logger.info(
        f"DESACTIVAR Usuario id={usuario.id}, por={request.user.username}"
    )

    return JsonResponse({'status': 'ok', 'message': 'Usuario desactivado.'})



@login_required
@csrf_exempt
def reactivar_usuario(request, user_id):
    if not _es_admin(request.user):
        return HttpResponseForbidden("Solo Administrador.")
    usuario = get_object_or_404(Usuario, id=user_id)

    usuario.estado = 'activo'
    usuario.activo = True
    usuario.save(update_fields=['estado', 'activo'])

    audit_logger.info(
        f"REACTIVAR Usuario id={usuario.id}, por={request.user.username}"
    )

    return JsonResponse({'status': 'ok', 'message': 'Usuario reactivado.'})



# ================================================================
# 🔥 Bloquear / desbloquear + Auditoría
# ================================================================

@login_required
@csrf_exempt
def bloquear_usuario(request, user_id):
    if not _es_admin(request.user):
        return HttpResponseForbidden("Solo Administrador.")

    usuario = get_object_or_404(Usuario, id=user_id)
    usuario.estado = 'bloqueado'
    usuario.activo = False
    usuario.save(update_fields=['estado', 'activo'])

    audit_logger.info(
        f"BLOQUEAR Usuario id={usuario.id}, por={request.user.username}"
    )

    return JsonResponse({'status': 'ok', 'message': 'Usuario bloqueado.'})



@login_required
@csrf_exempt
def desbloquear_usuario(request, user_id):
    if not _es_admin(request.user):
        return HttpResponseForbidden("Solo Administrador.")

    usuario = get_object_or_404(Usuario, id=user_id)
    usuario.estado = 'activo'
    usuario.activo = True
    usuario.save(update_fields=['estado', 'activo'])

    audit_logger.info(
        f"DESBLOQUEAR Usuario id={usuario.id}, por={request.user.username}"
    )

    return JsonResponse({'status': 'ok', 'message': 'Usuario desbloqueado.'})



# ================================================================
# 🔥 Reinicio de clave + Auditoría
# ================================================================

@login_required
@csrf_exempt
def reiniciar_clave(request: HttpRequest, user_id: int):
    if not _es_admin(request.user):
        return HttpResponseForbidden("Solo Administrador.")
    usuario = get_object_or_404(Usuario, id=user_id)

    invite_user_and_email(usuario, source='reset')

    audit_logger.info(
        f"REINICIAR_CLAVE Usuario id={usuario.id}, por={request.user.username}"
    )

    return JsonResponse({
        'status': 'ok',
        'message': f'Se envió un correo para reiniciar la clave a {usuario.email}.'
    })
