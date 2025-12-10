from datetime import datetime
import json
from decimal import Decimal, InvalidOperation

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.views.decorators.http import require_POST
from django.core.exceptions import ValidationError

from lilis_erp.roles import require_roles

from .models import MovimientoInventario, Producto, Proveedor, Bodega
from apps.api.serializers import (
    UsuarioSerializer,
    ProductoSerializer,
    ProveedorSerializer,
    MovimientoInventarioSerializer,
)

# Excel
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


# ==============================================================
#               BÚSQUEDA GLOBAL
# ==============================================================
def _build_transaction_q(q: str) -> Q:
    q = (q or "").strip()
    if not q:
        return Q()

    expr = (
        Q(producto__sku__icontains=q) |
        Q(producto__nombre__icontains=q) |
        Q(tipo__icontains=q) |
        Q(bodega_origen__nombre__icontains=q) |
        Q(bodega_destino__nombre__icontains=q) |
        Q(proveedor__razon_social__icontains=q) |
        Q(proveedor__rut_nif__icontains=q) |
        Q(creado_por__username__icontains=q) |
        Q(lote__icontains=q)
    )

    try:
        expr |= Q(cantidad=Decimal(str(q.replace(",", "."))))
    except:
        pass

    try:
        expr |= Q(id=int(q))
    except:
        pass

    return expr


# ==============================================================
#               LISTADO TRANSACCIONES
# ==============================================================
@login_required
@require_roles("ADMIN", "PRODUCCION", "INVENTARIO", "VENTAS", "COMPRAS")
def gestion_transacciones(request):

    query = request.GET.get("q", "")
    sort_by = request.GET.get("sort", "-id")
    ver = request.GET.get("ver", "todos")
    export = request.GET.get("export", "")

    valid_sort_fields = [
        "id", "-id", "fecha", "-fecha",
        "producto__nombre", "-producto__nombre", "tipo", "-tipo"
    ]

    if sort_by not in valid_sort_fields:
        sort_by = "-id"

    qs = MovimientoInventario.objects.select_related(
        "producto", "proveedor", "bodega_origen",
        "bodega_destino", "creado_por"
    )

    filtro_tipos = {
        "ingreso": "INGRESO",
        "salida": "SALIDA",
        "ajuste": "AJUSTE",
        "devolucion": "DEVOLUCION",
        "transferencia": "TRANSFERENCIA",
    }
    if ver in filtro_tipos:
        qs = qs.filter(tipo=filtro_tipos[ver])

    if query:
        qs = qs.filter(_build_transaction_q(query))

    qs = qs.order_by(sort_by)

    # --- Exportar Excel ---
    if export == "xlsx":
        if Workbook is None:
            return HttpResponse("Debes instalar openpyxl", status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos"

        headers = [
            "ID", "Fecha", "Tipo", "Producto", "SKU", "Cantidad",
            "Bodega Origen", "Bodega Destino", "Proveedor",
            "Lote", "Serie", "Vencimiento", "Usuario", "Observación",
        ]
        ws.append(headers)

        for m in qs:
            ws.append([
                m.id,
                m.fecha.strftime("%Y-%m-%d %H:%M"),
                m.tipo,
                m.producto.nombre,
                m.producto.sku,
                m.cantidad,
                m.bodega_origen.nombre if m.bodega_origen else "-",
                m.bodega_destino.nombre if m.bodega_destino else "-",
                m.proveedor.razon_social if m.proveedor else "-",
                m.lote or "-",
                m.serie or "-",
                m.fecha_vencimiento.strftime("%Y-%m-%d") if m.fecha_vencimiento else "-",
                m.creado_por.username if m.creado_por else "-",
                m.observacion or "",
            ])

        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        filename = f"movimientos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    # Paginador corregido
    paginator = Paginator(qs, 10)
    page_number = request.GET.get("page", 1)

    try:
        page_obj = paginator.page(page_number)
    except:
        page_obj = paginator.page(1)

    return render(request, "gestion_transacciones.html", {
        "movimientos": page_obj.object_list,
        "page_obj": page_obj,
        "query": query,
        "sort_by": sort_by,
        "ver": ver,
        "bodegas": Bodega.objects.all(),
    })


# ==============================================================
#               CREAR TRANSACCIÓN
# ==============================================================
@login_required
@require_roles("ADMIN", "PRODUCCION", "INVENTARIO")
@require_POST
def crear_transaccion(request):

    try:
        data = json.loads(request.body.decode("utf-8"))
    except:
        return JsonResponse({"ok": False, "error": "JSON inválido"}, status=400)

    errors = {}
    tipo = (data.get("tipo") or "").upper()
    producto_text = (data.get("producto_text") or "").strip()
    proveedor_text = (data.get("proveedor_text") or "").strip()
    cantidad_raw = data.get("cantidad")

    tipos_validos = {
        "INGRESO": "INGRESO",
        "SALIDA": "SALIDA",
        "AJUSTE": "AJUSTE",
        "DEVOLUCION": "DEVOLUCION",
        "DEVOLUCIÓN": "DEVOLUCION",
        "TRANSFERENCIA": "TRANSFERENCIA",
    }

    if tipo not in tipos_validos:
        return JsonResponse({"ok": False, "error": "Tipo inválido"}, status=400)

    tipo = tipos_validos[tipo]

    # Validar cantidad -> Decimal
    try:
        cantidad = Decimal(str(cantidad_raw))
        # Para ajustes, se permite 0. Para otros tipos, debe ser > 0.
        # Si es TRANSFERENCIA, ignoramos la validación de cantidad aquí porque se tomará de la BD.
        if tipo == "TRANSFERENCIA":
            pass
        elif tipo == "AJUSTE" and cantidad < 0:
            errors["cantidad"] = "La cantidad para un ajuste no puede ser negativa."
        elif tipo != "AJUSTE" and cantidad <= 0:
            errors["cantidad"] = "La cantidad debe ser mayor que cero."
    except (InvalidOperation, TypeError):
        # Si es transferencia, permitimos cantidad inválida/vacía temporalmente
        if tipo == "TRANSFERENCIA":
            cantidad = Decimal("0")
        else:
            errors["cantidad"] = "Cantidad inválida."

    if not producto_text:
        errors["producto_text"] = "Producto requerido."

    if tipo == "INGRESO" and not proveedor_text:
        errors["proveedor_text"] = "Proveedor requerido."

    if errors:
        return JsonResponse({"ok": False, "errors": errors}, status=400)

    # Producto
    producto = (
        Producto.objects.filter(sku=producto_text).first()
        or Producto.objects.filter(nombre__iexact=producto_text).first()
    )
    if not producto:
        return JsonResponse({"ok": False, "errors": {"producto_text": "Producto no encontrado"}}, status=400)

    # Proveedor
    proveedor = None
    if proveedor_text:
        proveedor = (
            Proveedor.objects.filter(rut_nif=proveedor_text).first()
            or Proveedor.objects.filter(razon_social__iexact=proveedor_text).first()
        )

    # Bodegas
    bodega_origen = None
    bodega_destino = None

    # Validamos que venga el ID y no sea nulo o vacío
    if data.get("bodega_origen") and str(data["bodega_origen"]).strip() not in ["", "null", "None"]:
        bodega_origen = Bodega.objects.filter(id=data["bodega_origen"]).first()

    # Validamos que venga el ID y no sea nulo o vacío
    if data.get("bodega_destino") and str(data["bodega_destino"]).strip() not in ["", "null", "None"]:
        bodega_destino = Bodega.objects.filter(id=data["bodega_destino"]).first()

    try:
        with transaction.atomic():

            mov = MovimientoInventario(
                tipo=tipo,
                producto=producto,
                proveedor=proveedor,
                cantidad=cantidad,
                lote=data.get("lote") or "",
                serie=data.get("serie") or "",
                fecha_vencimiento=data.get("vencimiento") or None,
                observacion=data.get("observaciones") or "",
                creado_por=request.user,
                bodega_origen=bodega_origen,
                bodega_destino=bodega_destino,
            )

            # --- LÓGICA DE ASIGNACIÓN DE BODEGAS ---
            
            # 1. Obtener bodega por defecto si falta alguna obligatoria
            default_bodega = Bodega.objects.order_by("id").first()
            if not default_bodega:
                raise ValidationError("No hay bodegas creadas en el sistema.")

            # 2. Asegurar bodegas según el tipo de movimiento
            if mov.tipo in ("INGRESO", "DEVOLUCION", "AJUSTE"):
                # Estos tipos requieren un DESTINO (donde entra o se ajusta el stock)
                if not mov.bodega_destino:
                    mov.bodega_destino = default_bodega
            
            elif mov.tipo == "SALIDA":
                # Salida requiere un ORIGEN
                if not mov.bodega_origen:
                    mov.bodega_origen = default_bodega

            elif mov.tipo == "TRANSFERENCIA":
                # Transferencia requiere AMBAS
                if not mov.bodega_origen:
                     # Intentar buscar una con stock si no se especificó
                    stock_origen = mov.producto.stocks.filter(cantidad__gt=0).order_by('-cantidad').first()
                    if stock_origen:
                        mov.bodega_origen = stock_origen.bodega
                    else:
                        mov.bodega_origen = default_bodega # Fallback
                
                if not mov.bodega_destino:
                    # Buscar una destino diferente a la origen
                    mov.bodega_destino = Bodega.objects.exclude(id=mov.bodega_origen.id).first()

                if not mov.bodega_destino:
                     raise ValidationError("Se requieren al menos 2 bodegas para realizar una transferencia.")
                
                if mov.bodega_origen.id == mov.bodega_destino.id:
                    raise ValidationError("La bodega de origen y destino no pueden ser la misma.")

            # Lógica solicitada: Si es TRANSFERENCIA, la cantidad es la que está en base de datos (Stock total de la bodega origen)
            if mov.tipo == "TRANSFERENCIA":
                # Buscamos el stock en la bodega de origen
                stock_item = mov.producto.stocks.filter(bodega=mov.bodega_origen).first()
                stock_actual = stock_item.cantidad if stock_item else Decimal("0")
                
                if stock_actual <= 0:
                    raise ValidationError(f"No hay stock disponible del producto '{mov.producto.nombre}' en la bodega '{mov.bodega_origen.nombre}' para transferir.")
                
                mov.cantidad = stock_actual

            mov.full_clean()
            mov.save()
            mov.aplicar_a_stock()

            return JsonResponse({"ok": True, "id": mov.id})

    except ValidationError as e:
        # Captura errores de validación tanto de `full_clean` como de `aplicar_a_stock`
        # y los devuelve como un error global para que el usuario los vea.
        error_message = "; ".join(getattr(e, 'messages', [str(e)]))
        return JsonResponse({"ok": False, "errors": {"__all__": error_message}}, status=400)

    except Exception as e:
        # Para cualquier otro error inesperado, devuelve un mensaje claro.
        error_message = f"Error inesperado: {str(e)}"
        return JsonResponse({"ok": False, "errors": {"__all__": error_message}}, status=500)


# ==============================================================
#               EDITAR TRANSACCIÓN
# ==============================================================
@login_required
@require_roles("ADMIN", "PRODUCCION", "INVENTARIO")
def editar_transaccion(request, mov_id):

    try:
        mov = MovimientoInventario.objects.get(id=mov_id)
    except MovimientoInventario.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Movimiento no encontrado"}, status=404)

    if request.method == "GET":
        return JsonResponse({
            "ok": True,
            "movimiento": {
                "id": mov.id,
                "fecha": mov.fecha.isoformat(),
                "tipo": mov.tipo,
                "producto": mov.producto.id,
                "proveedor": mov.proveedor.id if mov.proveedor else None,
                "cantidad": mov.cantidad,
            }
        })

    return JsonResponse({"ok": False, "error": "Editar transacciones no permitido"}, status=400)


# ==============================================================
#               ELIMINAR TRANSACCIÓN
# ==============================================================
@login_required
@require_roles("ADMIN", "PRODUCCION", "INVENTARIO")
@require_POST
def eliminar_transaccion(request, mov_id):
    return JsonResponse({"ok": False, "error": "No se pueden eliminar transacciones"}, status=400)
